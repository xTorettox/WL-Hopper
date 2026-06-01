import os
import shutil
import zipfile
import re
from io import BytesIO
from datetime import datetime, timedelta
import threading

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from PIL import Image

# Importaciones de tu scraper y utilidades actuales
from scraper import WLHopperBot, BureauVeritasBot, MicrosoftSharePointBot
from utils import extraer_internos, extraer_texto_de_archivo

# --- CONFIGURACIÓN DE TEMA ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("green")
VERDE_SULLAIR = "#008657"

class WLHopperDesktop(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("WL Hopper - Sullair Argentina (Desktop Edition)")
        self.geometry("1100x780")
        
        # Variables de estado (Equivalentes al session_state de Streamlit)
        self.res_lista = []
        self.df_excel = None
        self.hay_archivos = False
        self.ruta_temp = "descargas_temp"
        
        self.crear_interfaz()
        
    def log(self, mensaje):
        self.terminal.insert("end", f"{mensaje}\n")
        self.terminal.see("end")
        
    def crear_interfaz(self):
        # Configuración de Grid Principal (Izquierda: Panel, Derecha: Terminal)
        self.grid_columnconfigure(0, weight=10)
        self.grid_columnconfigure(1, weight=12)
        self.grid_rowconfigure(0, weight=1)
        
        # ---------------- COLUMNA IZQUIERDA (Panel de Control) ----------------
        self.frame_izq = ctk.CTkFrame(self, corner_radius=0)
        self.frame_izq.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Títulos principales
        self.lbl_titulo = ctk.CTkLabel(self.frame_izq, text="🚀 WL Hopper", font=ctk.CTkFont(size=22, weight="bold"))
        self.lbl_titulo.pack(pady=(15, 2))
        self.lbl_subtitulo = ctk.CTkLabel(self.frame_izq, text="Automatización de Certificados", text_color="gray", font=ctk.CTkFont(size=12))
        self.lbl_subtitulo.pack(pady=(0, 10))
        
        # CONTENEDOR DE CREDENCIALES
        self.frame_creds = ctk.CTkFrame(self.frame_izq)
        self.frame_creds.pack(fill="x", padx=15, pady=10)
        
        # Título del contenedor de credenciales
        self.lbl_creds_title = ctk.CTkLabel(self.frame_creds, text="🔐 Credenciales (Entorno Local)", font=ctk.CTkFont(weight="bold"))
        self.lbl_creds_title.grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 10), sticky="w")
        
        # Worklift Creds
        ctk.CTkLabel(self.frame_creds, text="Usuario WL:").grid(row=1, column=0, padx=(15, 5), pady=5, sticky="w")
        self.txt_wl_user = ctk.CTkEntry(self.frame_creds, width=160)
        self.txt_wl_user.grid(row=1, column=1, padx=(5, 15), pady=5, sticky="ew")
        self.txt_wl_user.insert(0, "etolosa@sullair.com.ar")
        
        ctk.CTkLabel(self.frame_creds, text="Contraseña WL:").grid(row=2, column=0, padx=(15, 5), pady=5, sticky="w")
        self.txt_wl_pass = ctk.CTkEntry(self.frame_creds, show="*", width=160)
        self.txt_wl_pass.grid(row=2, column=1, padx=(5, 15), pady=5, sticky="ew")
        self.txt_wl_pass.insert(0, "Sullair2025")
        
        # Bureau Veritas Creds
        ctk.CTkLabel(self.frame_creds, text="Usuario BV:").grid(row=3, column=0, padx=(15, 5), pady=5, sticky="w")
        self.txt_bv_user = ctk.CTkEntry(self.frame_creds, width=160)
        self.txt_bv_user.grid(row=3, column=1, padx=(5, 15), pady=5, sticky="ew")
        self.txt_bv_user.insert(0, "SULLAIRNQN")
        
        ctk.CTkLabel(self.frame_creds, text="Contraseña BV:").grid(row=4, column=0, padx=(15, 5), pady=5, sticky="w")
        self.txt_bv_pass = ctk.CTkEntry(self.frame_creds, show="*", width=160)
        self.txt_bv_pass.grid(row=4, column=1, padx=(5, 15), pady=5, sticky="ew")
        self.txt_bv_pass.insert(0, "SULLAIRNQN2024")
        
        # Microsoft SharePoint Creds
        ctk.CTkLabel(self.frame_creds, text="Usuario MS:").grid(row=5, column=0, padx=(15, 5), pady=5, sticky="w")
        self.txt_ms_user = ctk.CTkEntry(self.frame_creds, width=160)
        self.txt_ms_user.grid(row=5, column=1, padx=(5, 15), pady=5, sticky="ew")
        self.txt_ms_user.insert(0, "fcendra@sullair.com.ar")
        
        ctk.CTkLabel(self.frame_creds, text="Contraseña MS:").grid(row=6, column=0, padx=(15, 5), pady=(5, 15), sticky="w")
        self.txt_ms_pass = ctk.CTkEntry(self.frame_creds, show="*", width=160)
        self.txt_ms_pass.grid(row=6, column=1, padx=(5, 15), pady=(5, 15), sticky="ew")
        self.txt_ms_pass.insert(0, "C4n1ch3r1426")
        
        # Hacer que la columna de los inputs se expanda prolija
        self.frame_creds.grid_columnconfigure(1, weight=1)
        
        # OPCIONES DE DESCARGA
        self.frame_opciones = ctk.CTkFrame(self.frame_izq)
        self.frame_opciones.pack(fill="x", padx=15, pady=10)
        
        self.chk_cert = ctk.CTkCheckBox(self.frame_opciones, text="Descargar Certificados", text_color="white")
        self.chk_cert.select()
        self.chk_cert.grid(row=0, column=0, padx=10, pady=8, sticky="w")
        
        self.chk_inf = ctk.CTkCheckBox(self.frame_opciones, text="Descargar Informes", text_color="white")
        self.chk_inf.grid(row=0, column=1, padx=10, pady=8, sticky="w")
        
        self.chk_semestral = ctk.CTkCheckBox(self.frame_opciones, text="Vencimiento Semestral (180 días)", text_color="white")
        self.chk_semestral.grid(row=1, column=0, columnspan=2, padx=10, pady=8, sticky="w")
        
        self.chk_doc_equipo = ctk.CTkCheckBox(self.frame_opciones, text="Descargar Doc. Equipo (SharePoint)", text_color="white")
        self.chk_doc_equipo.grid(row=2, column=0, columnspan=2, padx=10, pady=8, sticky="w")
        
        # ENTRADA DE INTERNOS
        ctk.CTkLabel(self.frame_izq, text="Listado de Internos:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=15, pady=(10,0))
        self.txt_internos = ctk.CTkTextbox(self.frame_izq, height=100)
        self.txt_internos.pack(fill="x", padx=15, pady=5)
        self.txt_internos.insert("1.0", "E040230, 3797")
        
        # BOTÓN PRINCIPAL
        self.btn_run = ctk.CTkButton(self.frame_izq, text="🚀 COMENZAR PROCESO", fg_color=VERDE_SULLAIR, font=ctk.CTkFont(weight="bold"), command=self.start_proceso_thread)
        self.btn_run.pack(fill="x", padx=15, pady=15)
        
        # ---------------- COLUMNA DERECHA (Terminal de Actividad) ----------------
        self.frame_der = ctk.CTkFrame(self, corner_radius=0)
        self.frame_der.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_der, text=">_ REGISTRO DE ACTIVIDAD", font=ctk.CTkFont(family="Consolas", size=14, weight="bold"), text_color="#00ff00").pack(anchor="w", padx=15, pady=10)
        
        self.terminal = ctk.CTkTextbox(self.frame_der, font=ctk.CTkFont(family="Consolas", size=12), text_color="#f8f9fa", fg_color="#212529", border_color="#444", border_width=1)
        self.terminal.pack(fill="both", expand=True, padx=15, pady=10)
        
        # BOTONES DE DESCARGA LOCAL
        self.frame_descargas = ctk.CTkFrame(self.frame_der, fg_color="transparent")
        self.frame_descargas.pack(fill="x", padx=15, pady=10)
        
        self.btn_excel = ctk.CTkButton(self.frame_descargas, text="📊 Guardar Excel", state="disabled", command=self.guardar_excel)
        self.btn_excel.grid(row=0, column=0, padx=5, sticky="ew")
        
        self.btn_zip = ctk.CTkButton(self.frame_descargas, text="📂 Guardar ZIP", state="disabled", command=self.guardar_zip)
        self.btn_zip.grid(row=0, column=1, padx=5, sticky="ew")
        self.frame_descargas.grid_columnconfigure((0,1), weight=1)
 
    def set_inputs_state(self, state):
        self.txt_wl_user.configure(state=state)
        self.txt_wl_pass.configure(state=state)
        self.txt_bv_user.configure(state=state)
        self.txt_bv_pass.configure(state=state)
        self.txt_ms_user.configure(state=state)
        self.txt_ms_pass.configure(state=state)
        self.chk_cert.configure(state=state)
        self.chk_inf.configure(state=state)
        self.chk_semestral.configure(state=state)
        self.chk_doc_equipo.configure(state=state)
        self.txt_internos.configure(state=state)

    def start_proceso_thread(self):
        threading.Thread(target=self.ejecutar_proceso, daemon=True).start()

    def ejecutar_proceso(self):
        wl_usr = self.txt_wl_user.get().strip() or "etolosa@sullair.com.ar"
        wl_pass = self.txt_wl_pass.get().strip() or "Sullair2025"
        bv_usr = self.txt_bv_user.get().strip() or "SULLAIRNQN"
        bv_pass = self.txt_bv_pass.get().strip() or "SULLAIRNQN2024"
        ms_usr = self.txt_ms_user.get().strip() or "fcendra@sullair.com.ar"
        ms_pass = self.txt_ms_pass.get().strip() or "C4n1ch3r1426"
        
        if not wl_usr or not wl_pass:
            messagebox.showerror("Error", "Faltan las credenciales obligatorias de Worklift.")
            return
            
        bajar_doc_equipo = bool(self.chk_doc_equipo.get())
        if bajar_doc_equipo and (not ms_usr or not ms_pass):
            messagebox.showerror("Error", "Faltan las credenciales de Microsoft para SharePoint.")
            return
            
        self.btn_run.configure(state="disabled")
        self.btn_excel.configure(state="disabled")
        self.btn_zip.configure(state="disabled")
        self.set_inputs_state("disabled")
        self.terminal.delete("1.0", "end")
        
        try:
            if os.path.exists(self.ruta_temp): 
                shutil.rmtree(self.ruta_temp)
            os.makedirs(self.ruta_temp, exist_ok=True)
            
            bajar_cert = bool(self.chk_cert.get())
            bajar_inf = bool(self.chk_inf.get())
            es_semestral = bool(self.chk_semestral.get())
            
            texto_raw = self.txt_internos.get("1.0", "end")
            lista_internos = extraer_internos(texto_raw)
            
            if not lista_internos:
                self.log("❌ No se encontraron internos válidos para procesar.")
                return
                
            self.log("--- PARÁMETROS ---")
            self.log(f"Descargar certificados: {'Sí' if bajar_cert else 'No'}")
            self.log(f"Descargar informes de inspección: {'Sí' if bajar_inf else 'No'}")
            self.log(f"Descargar doc. equipo (SharePoint): {'Sí' if bajar_doc_equipo else 'No'}")
            self.log(f"Modo vencimiento semestral: {'Sí' if es_semestral else 'No'}")
            self.log("------------------")
            self.log("Iniciando conexión con Worklift...")
            
            bot = WLHopperBot(headless=False) 
            
            if bot.iniciar(wl_usr, wl_pass):
                self.log("🔐 Login exitoso en Worklift.")
                
                exito_bv_login = False
                if bv_usr and bv_pass:
                    self.log("Iniciando conexión con BV...")
                    bv_test_bot = BureauVeritasBot(headless=True)
                    exito_bv_login, error_bv_login = bv_test_bot.iniciar(bv_usr, bv_pass, pw_instance=bot.pw)
                    bv_test_bot.cerrar()
                    if exito_bv_login:
                        self.log("🔐 Login exitoso en BV.")
                    else:
                        self.log(f"❌ Falló conexión con BV: {error_bv_login}")
                        
                exito_ms_login = False
                ms_bot = None
                if bajar_doc_equipo and ms_usr and ms_pass:
                    self.log("Iniciando conexión con Microsoft SharePoint...")
                    ms_bot = MicrosoftSharePointBot(headless=True)
                    exito_ms_login, error_ms_login = ms_bot.iniciar(ms_usr, ms_pass, pw_instance=bot.pw)
                    if exito_ms_login:
                        self.log("🔐 Login exitoso en SharePoint.")
                    else:
                        self.log(f"❌ Falló conexión con SharePoint: {error_ms_login}")
                
                self.res_lista = []
                for int_id in lista_internos:
                    self.log(f"\n--- Procesando {int_id} ---")
                    res = bot.procesar_interno(int_id, self.ruta_temp, bajar_cert, bajar_inf, es_semestral=es_semestral, prefijo_cert="")
                    res['id'] = int_id
                    res['proveedor'] = "Worklift"
                    
                    if bv_usr and bv_pass and exito_bv_login:
                        bv_bot = BureauVeritasBot(headless=True)
                        exito_bv, error_bv = bv_bot.iniciar(bv_usr, bv_pass, pw_instance=bot.pw)
                        
                        if exito_bv:
                            bv_res = bv_bot.procesar_interno(int_id, self.ruta_temp, bajar_cert=bajar_cert, bajar_inf=bajar_inf, prefijo_cert="")
                            bv_bot.cerrar()
                            
                            # Si BV encontró algo útil, comparamos con WL
                            if bv_res.get('descargado') or bv_res.get('status') == 'VIGENTE (BV)' or bv_res.get('status') == 'Encontrado en BV':
                                
                                # Funciones helper para fechas
                                def get_dt(d_str):
                                    if not d_str or d_str == "-": return datetime.min
                                    try: return datetime.strptime(d_str, "%d/%m/%Y")
                                    except: return datetime.min
                                    
                                wl_v_dt = get_dt(res.get('venc_real') if 'venc_real' in res else res.get('venc'))
                                bv_v_dt = get_dt(bv_res.get('venc'))
                                
                                wl_i_dt = get_dt(res.get('insp'))
                                bv_i_dt = get_dt(bv_res.get('insp'))
                                
                                hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                                
                                bv_tiene_cert_vigente = (bv_v_dt >= hoy)
                                wl_tiene_cert_vigente = (wl_v_dt >= hoy)
                                
                                gana_bv_cert = False
                                gana_bv_inf = False
                                
                                # DECISIÓN CERTIFICADO
                                if bv_tiene_cert_vigente and (bv_v_dt > wl_v_dt):
                                    gana_bv_cert = True
                                    
                                # DECISIÓN INFORME
                                if gana_bv_cert:
                                    gana_bv_inf = True
                                else:
                                    if bv_i_dt > wl_i_dt:
                                        gana_bv_inf = True
                                        
                                if gana_bv_cert or gana_bv_inf:
                                    res['proveedor'] = "Bureau Veritas"
                                    res['cert'] = bv_res.get('cert', 'NO') if gana_bv_cert else res.get('cert', 'NO')
                                    res['inf'] = bv_res.get('informe', 'NO') if gana_bv_inf else res.get('inf', 'NO')
                                    
                                    if gana_bv_inf:
                                        res['insp'] = bv_res.get('insp', res.get('insp'))
                                        
                                    res['venc_real'] = bv_res.get('venc') if gana_bv_cert else res.get('venc_real', res.get('venc'))
                                    
                                    if es_semestral:
                                        i_dt = get_dt(res['insp'])
                                        if i_dt != datetime.min:
                                            res['venc'] = (i_dt + timedelta(days=180)).strftime("%d/%m/%Y")
                                        else:
                                            res['venc'] = "-"
                                    else:
                                        if gana_bv_cert:
                                            res['venc'] = bv_res.get('venc', res.get('venc'))
                                            
                                    dias_restantes = (get_dt(res['venc']) - hoy).days if res['venc'] != "-" else -1
                                    
                                    if dias_restantes > 30:
                                        res['status'] = "VIGENTE"
                                        res['color'] = "VERDE"
                                        res['obs_final'] = f"{dias_restantes} días de vigencia."
                                        res['accion_final'] = "-"
                                    elif 0 <= dias_restantes <= 30:
                                        res['status'] = "PRÓXIMO A VENCER"
                                        res['color'] = "AMARILLO"
                                        res['obs_final'] = f"{dias_restantes} días de vigencia."
                                        res['accion_final'] = "Coordinar recertificación"
                                    else:
                                        res['status'] = "VENCIDO"
                                        res['color'] = "ROJO"
                                        if es_semestral:
                                            res['obs_final'] = f"Último certificado vencido en {res['venc']}." if res['venc'] != "-" else "Último certificado vencido."
                                        else:
                                            res['obs_final'] = "Último certificado vencido."
                                        obs_bv = bv_res.get('observaciones', '')
                                        if obs_bv and gana_bv_inf: res['obs_final'] += f"\nObservaciones BV: {obs_bv}"
                                        res['accion_final'] = "Coordinar recertificación urgente"
                                        
                                    res['log'] = [
                                        f"📄 Último Informe de Inspección: {res['insp']} (BV)",
                                        f"📅 Fecha vencimiento certificado: {res['venc']} (BV)"
                                    ]
                                    if dias_restantes <= 30:
                                        res['log'].append(f"💡 Sugerencia: {res['accion_final']}")
                                        
                                    # Eliminar archivos perdedores WL
                                    archivos_wl = res.get("archivos_descargados", [])
                                    for f_path in archivos_wl:
                                        if os.path.exists(f_path):
                                            if ("Certificado" in f_path and gana_bv_cert) or ("Informe" in f_path and gana_bv_inf):
                                                try: os.remove(f_path)
                                                except: pass
                                                
                                    # Eliminar archivos perdedores BV
                                    archivos_bv = bv_res.get("archivos_descargados", [])
                                    for f_path in archivos_bv:
                                        if os.path.exists(f_path):
                                            if ("Certificado" in f_path and not gana_bv_cert) or ("Informe" in f_path and not gana_bv_inf):
                                                try: os.remove(f_path)
                                                except: pass
                                else:
                                    # Ganó WL. Borramos los que bajó BV
                                    archivos_bv = bv_res.get("archivos_descargados", [])
                                    for f_path in archivos_bv:
                                        if os.path.exists(f_path):
                                            try: os.remove(f_path)
                                            except: pass
                                            
                    # --- Microsoft SharePoint Integration ---
                    doc_equipo_archivo = "-"
                    doc_equipo_tipo = "-"
                    if bajar_doc_equipo and exito_ms_login and ms_bot:
                        self.log(f"🔗 Consultando SharePoint para {int_id}...")
                        ms_res = ms_bot.procesar_interno(int_id, self.ruta_temp, prefijo_cert="")
                        for step in ms_res.get('log', []):
                            self.log(f"  [SharePoint] {step}")
                        if ms_res.get('descargado'):
                            doc_equipo_archivo = ms_res.get('archivo', '-')
                            doc_equipo_tipo = ms_res.get('tipo_doc', '-')
                            self.log(f"✅ Encontrado en SharePoint: {doc_equipo_tipo} ({doc_equipo_archivo})")
                        else:
                            self.log(f"❌ No se encontró documentación del equipo en SharePoint.")
                    res['doc_equipo_archivo'] = doc_equipo_archivo
                    res['doc_equipo_tipo'] = doc_equipo_tipo
                            
                    # --- IMPRESIÓN DEL LOG REESTRUCTURADO ---
                    self.log(f"Proveedor: {res.get('proveedor', 'Worklift')}")
                    
                    desc_inf = " (descargado)" if res.get('inf') == "SI" else ""
                    self.log(f"Última Inspección: {res.get('insp')}{desc_inf}")
                    
                    desc_cert = " (descargado)" if res.get('cert') == "SI" else ""
                    
                    color_emoji = "⚪"
                    color_code = res.get('color', '').upper()
                    if color_code == "VERDE": color_emoji = "🟢"
                    elif color_code == "AMARILLO": color_emoji = "🟡"
                    elif color_code == "ROJO": color_emoji = "🔴"
                    else:
                        st_upper = res.get('status', '').upper()
                        if "VIGENTE" in st_upper or "VERDE" in st_upper: color_emoji = "🟢"
                        elif "PRÓXIMO" in st_upper or "AMARILLO" in st_upper: color_emoji = "🟡"
                        elif "VENCIDO" in st_upper or "ROJO" in st_upper: color_emoji = "🔴"
                        
                    status_tag = f" ({color_emoji} {res.get('status', 'VIGENTE')})"
                    self.log(f"Último Certificado: {res.get('venc_real', res.get('venc'))}{desc_cert}{status_tag}")
                    
                    if doc_equipo_archivo != "-":
                        self.log(f"Doc. Equipo ({doc_equipo_tipo}): {doc_equipo_archivo}")
                    
                    # Cálculo de vigencia
                    def get_dt(d_str):
                        if not d_str or d_str == "-": return datetime.min
                        try: return datetime.strptime(d_str, "%d/%m/%Y")
                        except: return datetime.min
                    v_dt = get_dt(res.get('venc'))
                    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                    if v_dt != datetime.min:
                        dias = (v_dt - hoy).days
                        if dias >= 0:
                            dias_str = f"{dias} días de vigencia"
                        else:
                            dias_str = f"vencido hace {-dias} días"
                    else:
                        dias_str = "sin registro"
                    self.log(f"Días de vigencia: {dias_str}")
                    
                    if res.get('accion_final') != "-":
                        self.log(f"Comentarios: {res.get('accion_final')}")
                        
                    # Conservar advertencias OCR y sugerencias especiales
                    for log_line in res.get('log', []):
                        if any(log_line.strip().startswith(x) for x in ["⚠️", "🤖", "💡"]):
                            if log_line.strip().startswith("💡 Sugerencia:") and res.get('accion_final') != "-":
                                continue
                            self.log(f"  {log_line.strip()}")
                            
                    self.res_lista.append(res)
                    
                if ms_bot:
                    ms_bot.cerrar()
                bot.cerrar()
                self.log("\n🔓 Sesión cerrada correctamente.")
                self.log("🏁 PROCESO FINALIZADO.")
                
                self.hay_archivos = len(os.listdir(self.ruta_temp)) > 0
                self.btn_excel.configure(state="normal")
                if self.hay_archivos:
                    self.btn_zip.configure(state="normal")
            else:
                self.log("❌ ERROR: Credenciales de Worklift incorrectas.")
                messagebox.showerror("Error de Login", "No se pudo iniciar sesión en Worklift.")
        finally:
            self.btn_run.configure(state="normal")
            self.set_inputs_state("normal")

    def guardar_excel(self):
        import pandas as pd
        excel_data = []
        es_semestral = bool(self.chk_semestral.get())
        bajar_doc_equipo = bool(self.chk_doc_equipo.get())
        
        for r in self.res_lista:
            st_text = r.get("status", "N/A")
            row = {
                "INTERNO": r["id"],
                "PROVEEDOR": r.get("proveedor", "Worklift"),
                "ESTADO": st_text,
                "ÚLTIMA INSPECCIÓN": r.get("insp", "N/A"),
            }
            if es_semestral:
                row["VENCIMIENTO SEMESTRAL"] = r.get("venc", "N/A")
                row["VENCIMIENTO REAL"] = r.get("venc_real", r.get("venc", "N/A"))
            else:
                row["VENCIMIENTO ÚLTIMO CERTIFICADO"] = r.get("venc", "N/A")
                
            if "doc_equipo_archivo" in r:
                row["DOC. EQUIPO"] = r.get("doc_equipo_tipo", "-") if r.get("doc_equipo_archivo", "-") != "-" else "-"
                
            row.update({
                "CERTIFICADO": r.get("cert", "N/A"),
                "INFORME": r.get("inf", "N/A"),
                "OBSERVACIONES": r.get("obs_final", "-"),
                "ACCIONES": r.get("accion_final", "-")
            })
            excel_data.append(row)
            
        df = pd.DataFrame(excel_data)
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Reporte')
                worksheet = writer.sheets['Reporte']
                from openpyxl.styles import PatternFill, Font, Alignment
                
                # Ajustar ancho de columnas y alinear
                for column_cells in worksheet.columns:
                    col_letra = column_cells[0].column_letter
                    header_val = column_cells[0].value
                    
                    # Calcular longitud para auto-fit básico
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    ancho_final = min(length + 2, 50)
                    
                    if header_val == "ESTADO":
                        ancho_final = max(ancho_final, 22)
                    elif header_val == "DOC. EQUIPO":
                        ancho_final = 20
                    elif header_val == "OBSERVACIONES":
                        ancho_final = 19.57
                    elif header_val == "ACCIONES":
                        ancho_final = 27.43
                    elif header_val == "INFORME":
                        ancho_final = 9.5
                    
                    worksheet.column_dimensions[col_letra].width = ancho_final
                    
                    # Aplicar centrado y wrap a todas las celdas
                    for cell in column_cells:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Colores
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                green_font = Font(color="006100", bold=True)
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                yellow_font = Font(color="9C5700", bold=True)
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006", bold=True)
                
                header_fill = PatternFill(start_color="008657", end_color="008657", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                if es_semestral:
                    worksheet.insert_rows(1)
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(worksheet[2]))
                    title_cell = worksheet.cell(row=1, column=1)
                    title_cell.value = "⚠️ REPORTE DE VENCIMIENTOS SEMESTRALES (180 DÍAS)"
                    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                    title_cell.fill = header_fill
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    header_row = 2
                    worksheet.row_dimensions[1].height = 25
                else:
                    header_row = 1
                
                for cell in worksheet[header_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                headers = [cell.value for cell in worksheet[header_row]]
                estado_idx = headers.index("ESTADO") + 1 if "ESTADO" in headers else -1
                
                for row in range(header_row + 1, worksheet.max_row + 1):
                    res_idx = row - (header_row + 1)
                    color_code = ""
                    if res_idx >= 0 and res_idx < len(self.res_lista):
                        color_code = self.res_lista[res_idx].get('color', '').upper()
                    
                    for idx in [estado_idx]:
                        if idx != -1:
                            cell = worksheet.cell(row=row, column=idx)
                            if cell.value:
                                val = str(cell.value).upper()
                                if idx == estado_idx and color_code:
                                    if color_code == "VERDE":
                                        cell.fill = green_fill
                                        cell.font = green_font
                                    elif color_code == "AMARILLO":
                                        cell.fill = yellow_fill
                                        cell.font = yellow_font
                                    elif color_code == "ROJO":
                                        cell.fill = red_fill
                                        cell.font = red_font
                                else:
                                    if "VERDE" in val or "VIGENTE" in val or "APROBADO" in val:
                                        cell.fill = green_fill
                                        cell.font = green_font
                                    elif "AMARILLO" in val or "PRÓXIMO" in val or "GESTIÓN" in val:
                                        cell.fill = yellow_fill
                                        cell.font = yellow_font
                                    elif "ROJO" in val or "VENCIDO" in val or "RECHAZADO" in val:
                                        cell.fill = red_fill
                                        cell.font = red_font
                                        
            messagebox.showinfo("Guardado", f"Reporte Excel generado con formato en:\n{filepath}")

    def guardar_zip(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("ZIP files", "*.zip")])
        if filepath:
            with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
                for root, _, files in os.walk(self.ruta_temp):
                    for file in files:
                        zf.write(os.path.join(root, file), file)
            messagebox.showinfo("Guardado", f"Archivo comprimido creado en:\n{filepath}")

if __name__ == "__main__":
    app = WLHopperDesktop()
    app.mainloop()
