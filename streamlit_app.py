import streamlit as st
import os
import shutil
import pandas as pd
import zipfile
from io import BytesIO
from scraper import WLHopperBot, BureauVeritasBot
from utils import extraer_internos, extraer_texto_de_archivo, calcular_vencimiento_semestral, asegurar_carpeta
import streamlit.components.v1 as components
import pytesseract
from PIL import Image, ImageEnhance
import re
import importlib
import utils
importlib.reload(utils)
from utils import extraer_internos, extraer_texto_de_archivo, calcular_vencimiento_semestral, asegurar_carpeta

from supabase import create_client, Client
from cryptography.fernet import Fernet
from datetime import datetime, timedelta

# --- CONFIGURACIÓN BÚNKER DE SEGURIDAD Y BD ---
try:
    SUPABASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    FERNET_KEY = st.secrets["FERNET_KEY"].encode()
    cipher_suite = Fernet(FERNET_KEY)
except Exception as e:
    st.error(f"Error cargando secrets: {e}")
    supabase = None
    cipher_suite = None

def encriptar(texto):
    if not texto or not cipher_suite: return ""
    return cipher_suite.encrypt(texto.encode()).decode()

def desencriptar(texto_cifrado):
    if not texto_cifrado or not cipher_suite: return ""
    try:
        return cipher_suite.decrypt(texto_cifrado.encode()).decode()
    except:
        return ""

def get_location_info(ip):
    if not ip or ip == "Desconocido": return "Desconocido"
    try:
        ip = ip.split(",")[0].strip()
        import requests
        r = requests.get(f"http://ip-api.com/json/{ip}", timeout=2)
        if r.status_code == 200:
            data = r.json()
            if data.get("status") == "success":
                return f"{data.get('city', '')}, {data.get('regionName', '')}, {data.get('countryCode', '')}".strip(", ")
    except:
        pass
    return "Desconocido"

def registrar_metrica(interno, fuente, exito=True):
    if not supabase: return
    try:
        # Intenta obtener IP (En Streamlit Cloud suele venir en los headers)
        try:
            from streamlit.web.server.websocket_headers import _get_websocket_headers
            headers = _get_websocket_headers()
            ip = headers.get("X-Forwarded-For", "Desconocido") if headers else "Desconocido"
        except:
            ip = "Desconocido"
            
        ubicacion = get_location_info(ip)
        
        minutos = 1
        data = {
            "usuario": st.session_state.get("logged_user", "desconocido"),
            "equipo": interno,
            "fuente": fuente,
            "fecha": datetime.now().isoformat(),
            "minutos_ahorrados": minutos,
            "exito": exito,
            "ip": f"{ip} ({ubicacion})"
        }
        supabase.table("metricas").insert(data).execute()
    except Exception as e:
        print(f"Error registrando métrica: {e}")


# --- CONFIGURACIÓN ---
st.set_page_config(page_title="WL Hopper - Sullair Argentina", page_icon="img/favicon.png", layout="wide")

# --- ESTILOS CSS ---
    
VERDE_SULLAIR = "#008657"
st.markdown(f"""
    <style>
    /* Base de la terminal */
    .terminal-box {{
        background-color: #212529; color: #f8f9fa; font-family: 'Consolas', monospace;
        font-size: 13px; padding: 15px; border-radius: 5px; 
        overflow-y: auto; border: 1px solid #444;
    }}
    
    /* Layout exclusivo para Desktop (pantallas anchas) */
    @media (min-width: 768px) {{
        .terminal-box {{
            position: absolute; top: 0; bottom: 0; left: 0; right: 0; width: 100%; height: 100%; min-height: 535px;
        }}
        [data-testid="stHorizontalBlock"] {{
            align-items: stretch;
        }}
        [data-testid="stColumn"] {{
            position: relative;
        }}
    }}
    
    /* Layout para Móviles */
    @media (max-width: 767px) {{
        .terminal-box {{
            height: 400px;
            margin-top: 10px;
        }}
    }}
    
    /* Estilo para que el botón deshabilitado no flote */
    .stDownloadButton button {{
        margin-bottom: 0px !important;
        height: 45px !important;
    }}
    
    div.stButton > button:first-child {{ background-color: {VERDE_SULLAIR} !important; color: white !important; font-weight: bold; }}
    .log-entry {{ margin-bottom: 5px; border-bottom: 1px solid #333; padding-bottom: 2px; }}
    .logo-container {{ display: flex; justify-content: center; align-items: center; flex-direction: column; margin-bottom: 10px; }}
    
    @keyframes ellipsis {{
      0% {{ content: ""; }}
      25% {{ content: "."; }}
      50% {{ content: ".."; }}
      75% {{ content: "..."; }}
      100% {{ content: ""; }}
    }}
    .loading-dots::after {{
      content: "";
      animation: ellipsis 1.5s infinite;
    }}
    
    @keyframes blink {{
        0% {{ opacity: 1; }}
        50% {{ opacity: 0; }}
        100% {{ opacity: 1; }}
    }}
    .blink-cursor {{
        animation: blink 1s step-end infinite;
    }}
    </style>
    """, unsafe_allow_html=True)

# --- FUNCIÓN DE LOGIN (Control de Acceso) ---
def es_usuario_ingresos():
    """Devuelve True si el usuario actual pertenece al área de Ingresos."""
    username = st.session_state.get("logged_user", "")
    if not username: return False
    
    # Lista artesanal whitelisted
    if username in ["fcendra", "damian", "mkette", "mmette", "dkette", "dausili"]:
        return True
        
    # Intento de base de datos Supabase
    if supabase:
        try:
            res = supabase.table("usuarios_areas").select("area").eq("usuario", username).execute()
            if res.data and res.data[0].get("area", "").lower() == "ingresos":
                return True
        except:
            pass
            
    return False

def check_password():
    """Devuelve True si el usuario ingresó credenciales válidas."""
    def password_entered():
        # Verificamos contra los secrets de Streamlit
        if st.session_state["username"] in st.secrets["passwords"] and \
           st.session_state["password"] == st.secrets["passwords"][st.session_state["username"]]:
            st.session_state["password_correct"] = True
            st.session_state["logged_user"] = st.session_state["username"]
            del st.session_state["password"]  # Borramos la clave por seguridad
            
            # Fetch credenciales desde Supabase
            if supabase:
                try:
                    res = supabase.table("credenciales_sitios").select("*").eq("usuario_app", st.session_state["logged_user"]).execute()
                    wl_creds = {}
                    bv_creds = {}
                    ms_creds = {}
                    if res.data:
                        for cred in res.data:
                            s_name = cred.get("sitio", "")
                            u_dec = desencriptar(cred.get("user_enc", ""))
                            p_dec = desencriptar(cred.get("pass_enc", ""))
                            
                            if u_dec:
                                if s_name == "WL" or s_name.startswith("WL_"):
                                    wl_creds[u_dec] = p_dec
                                elif s_name.startswith("BV_"):
                                    bv_creds[u_dec] = p_dec
                                elif s_name.startswith("MS_"):
                                    ms_creds[u_dec] = p_dec
                    
                    st.session_state["wl_creds_dict"] = wl_creds
                    st.session_state["bv_creds_dict"] = bv_creds
                    st.session_state["ms_creds_dict"] = ms_creds
                    
                    if wl_creds:
                        first_wl = list(wl_creds.keys())[0]
                        st.session_state["wl_user"] = first_wl
                        st.session_state["wl_pw"] = wl_creds[first_wl]
                    else:
                        st.session_state["wl_user"], st.session_state["wl_pw"] = "", ""
                        
                    if bv_creds:
                        first_bv = list(bv_creds.keys())[0]
                        st.session_state["bv_user"] = first_bv
                        st.session_state["bv_pw"] = bv_creds[first_bv]
                    else:
                        st.session_state["bv_user"], st.session_state["bv_pw"] = "", ""

                    if ms_creds:
                        first_ms = list(ms_creds.keys())[0]
                        st.session_state["ms_user"] = first_ms
                        st.session_state["ms_pw"] = ms_creds[first_ms]
                    else:
                        st.session_state["ms_user"], st.session_state["ms_pw"] = "", ""
                except Exception as e:
                    print(f"Error fetching credentials: {e}")
        else:
            st.session_state["password_correct"] = False

    # Si NO es True (es decir, es False o None), mostramos el login
    if st.session_state.get("password_correct") is not True:
        c_l1, c_l2, c_l3 = st.columns([1.2, 1, 1.2])
        
        with c_l2:
            try:
                st.image("img/WL Hopper Logo - nspc.png", use_container_width=True)
            except:
                st.markdown("<h2 style='text-align: center;'>🚀 WL Hopper</h2>", unsafe_allow_html=True)
            
            st.markdown("<h4 style='text-align: center;'>Acceso al Sistema</h4>", unsafe_allow_html=True)
            
            with st.form("login_form"):
                st.text_input("Usuario", key="username")
                st.text_input("Contraseña", type="password", key="password")
                st.form_submit_button("Ingresar", on_click=password_entered, use_container_width=True)
            
            if st.session_state.get("password_correct") == False:
                st.error("😕 Usuario o contraseña incorrectos")
        
        return False
        
    return True

# --- FLUJO PRINCIPAL ---
if check_password():

    # Detección de móvil
    components.html("""
        <script>
        const isMobile = window.innerWidth < 768;
        window.parent.postMessage({type: 'streamlit:setComponentValue', value: isMobile}, '*');
        </script>
    """, height=0)

    # --- ADMIN DASHBOARD (SIDEBAR) ---
    if st.session_state.get("logged_user") == "fcendra" and supabase:
        with st.sidebar:
            st.markdown("### ⚙️ Panel de Admin")
            if st.button("📊 Abrir Dashboard", use_container_width=True):
                st.session_state.show_dashboard = True
                st.rerun()

    if st.session_state.get("show_dashboard", False):
        st.title("📊 Dashboard de Métricas - WL Hopper")
        if st.button("⬅️ Volver a la App", type="primary"):
            st.session_state.show_dashboard = False
            st.rerun()
            
        try:
            res_met = supabase.table("metricas").select("*").execute()
            if not res_met.data:
                st.info("No hay métricas registradas aún.")
            else:
                import pandas as pd
                df = pd.DataFrame(res_met.data)
                df['fecha'] = pd.to_datetime(df['fecha'])
                
                # --- FILTRO TEMPORAL DINÁMICO ---
                st.markdown("### Filtro Temporal")
                # Por defecto muestra los últimos 30 días o desde el primer registro disponible
                min_date = df['fecha'].min().date()
                max_date = df['fecha'].max().date()
                default_start = max(min_date, max_date - timedelta(days=30))
                rango = st.date_input("Seleccionar rango de fechas", value=(default_start, max_date), min_value=min_date, max_value=max_date)
                
                if isinstance(rango, tuple) and len(rango) == 2:
                    start_date, end_date = rango
                    df = df[(df['fecha'].dt.date >= start_date) & (df['fecha'].dt.date <= end_date)]
                
                # KPIs
                total_items = len(df)
                total_minutos = df['minutos_ahorrados'].sum() if 'minutos_ahorrados' in df.columns else 0
                horas_ahorradas = total_minutos / 60
                
                exitosos = df['exito'].sum() if 'exito' in df.columns else total_items
                tasa_exito = (exitosos / total_items * 100) if total_items > 0 else 0
                
                c1, c2, c3 = st.columns(3)
                c1.metric("🤖 Total Ejecuciones", total_items)
                c2.metric("⏳ Tiempo Ahorrado", f"{horas_ahorradas:.1f} hs")
                c3.metric("✅ Tasa de Efectividad", f"{tasa_exito:.1f}%")
                
                st.divider()
                
                c_graf1, c_graf2 = st.columns(2)
                with c_graf1:
                    st.markdown("##### 👥 Uso por Usuario")
                    uso_usuario = df['usuario'].value_counts()
                    st.bar_chart(uso_usuario)
                    
                with c_graf2:
                    st.markdown("##### 📈 Actividad en el Tiempo")
                    # Agrupar por fecha
                    df['dia'] = df['fecha'].dt.date
                    actividad_diaria = df['dia'].value_counts().sort_index()
                    st.line_chart(actividad_diaria)
                    
                st.markdown("##### 📝 Registros Detallados")
                # Ocultar algunas columnas internas o mostrarlas mejor formateadas
                st.dataframe(df.drop(columns=['dia']).sort_values(by='fecha', ascending=False), use_container_width=True)
                
                # Exportar Excel
                import io
                excel_buffer = io.BytesIO()
                df.drop(columns=['dia']).to_excel(excel_buffer, index=False)
                st.download_button("📥 Descargar Reporte Completo (Excel)", data=excel_buffer.getvalue(), file_name="Reporte_Metricas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"Error cargando el Dashboard: {e}")
            
        st.stop()

    @st.dialog("Acerca de WL Hopper")
    def mostrar_about():
        c1, c2 = st.columns([1, 2])
        with c1:
            try:
                st.image("img/robot_diapos.png")
            except:
                st.write("🤖")
        with c2:
            st.markdown("""
            **WL Hopper** es una app diseñada para optimizar la descarga de certificados desde los sitios de **Worklift** y **Bureau Veritas**.
            
            Inspirada en una tarea repetitiva que no quería seguirlo siendo, esta herramienta usa bots de navegación para descargar PDFs en segundo plano.
            """)
        
        st.info("🚀 **Misión:** Automatizar y acelerar la tarea de descarga masiva de certificados e informes, y recuperar y estructurar la información de nuestros equipos desde los sitios web de Worklift y Bureau Veritas.")
        
        st.divider()
        st.caption("Desarrollado por Fede García Cendra - 2026")

    with st.sidebar:
        st.markdown("### 🛠️ Opciones")
        st.button("Cerrar Sesión", on_click=lambda: st.session_state.clear(), use_container_width=True)
        if st.button("Acerca del Proyecto", use_container_width=True):
            mostrar_about()
    
    if "log_history" not in st.session_state: st.session_state.log_history = []
    if "proceso_completo" not in st.session_state: st.session_state.proceso_completo = False
    if "html_excel" not in st.session_state: st.session_state.html_excel = ""
    if "df_excel" not in st.session_state: st.session_state.df_excel = None
    if "hay_archivos" not in st.session_state: st.session_state.hay_archivos = False
    if "res_lista" not in st.session_state: st.session_state.res_lista = []
    if "texto_area" not in st.session_state: st.session_state.texto_area = ""
    if "ultimo_archivo_procesado" not in st.session_state: st.session_state.ultimo_archivo_procesado = None
    if "ejecutando" not in st.session_state: st.session_state.ejecutando = False
    is_exec = st.session_state.ejecutando

    
    st.markdown('<div class="logo-container">', unsafe_allow_html=True)
    c_l1, c_l2, c_l3 = st.columns([1.5, 1, 1.5])
    with c_l2: 
        try:
            st.image("img/WL Hopper Logo - nspc.png", use_container_width=True)
        except:
            pass
    st.markdown("<h5 style='text-align: center; color: #555; margin-top:-10px; margin-bottom: 25px;'>Automatización de Descarga de Certificados</h5>", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    with st.expander("🔐 Credenciales", expanded=not is_exec):
        st.write("Gestiona tus credenciales de forma cifrada en Supabase. Las claves quedan vinculadas a tu cuenta de usuario.")
        
        wl_creds_dict = st.session_state.get("wl_creds_dict", {})
        bv_creds_dict = st.session_state.get("bv_creds_dict", {})
        ms_creds_dict = st.session_state.get("ms_creds_dict", {})
        
        # --- GRILLA DE PERFILES CONFIGURADOS ---
        st.markdown("##### Perfiles Activos")
        
        # Determinar cuántas columnas necesitamos (2 o 3 según área)
        show_ms = es_usuario_ingresos()
        cols_count = 3 if show_ms else 2
        cols = st.columns(cols_count)
        
        # Column 1: Worklift
        with cols[0]:
            with st.container(border=True):
                st.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
                try:
                    st.image("img/WL-Logo.png", width=64)
                except:
                    st.markdown("<h3>WL</h3>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                
                wl_opciones = list(wl_creds_dict.keys())
                if wl_opciones:
                    sel_wl = st.selectbox("Perfil Worklift", wl_opciones, key="sel_wl_active", disabled=is_exec)
                    st.caption(f"Usuario: `{sel_wl}`")
                    
                    if st.button("🗑️ Eliminar WL", key="del_wl", disabled=is_exec, use_container_width=True):
                        if supabase:
                            sitio_val = f"WL_{st.session_state.get('logged_user', '')}_{sel_wl}"
                            try:
                                supabase.table("credenciales_sitios").delete().eq("sitio", sitio_val).execute()
                                del wl_creds_dict[sel_wl]
                                st.session_state["wl_creds_dict"] = wl_creds_dict
                                st.success("Eliminado")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
                else:
                    st.warning("Sin configurar")
                    sel_wl = ""
                    
        # Column 2: Bureau Veritas
        with cols[1]:
            with st.container(border=True):
                st.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
                try:
                    st.image("img/BV-Logo.png", width=64)
                except:
                    st.markdown("<h3>BV</h3>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                
                bv_opciones = list(bv_creds_dict.keys())
                if bv_opciones:
                    sel_bv = st.selectbox("Perfil BV", bv_opciones, key="sel_bv_active", disabled=is_exec)
                    st.caption(f"Usuario: `{sel_bv}`")
                    
                    if st.button("🗑️ Eliminar BV", key="del_bv", disabled=is_exec, use_container_width=True):
                        if supabase:
                            sitio_val = f"BV_{st.session_state.get('logged_user', '')}_{sel_bv}"
                            try:
                                supabase.table("credenciales_sitios").delete().eq("sitio", sitio_val).execute()
                                del bv_creds_dict[sel_bv]
                                st.session_state["bv_creds_dict"] = bv_creds_dict
                                st.success("Eliminado")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
                else:
                    st.warning("Sin configurar")
                    sel_bv = ""

        # Column 3: Microsoft SharePoint (Only if ingresos)
        if show_ms:
            with cols[2]:
                with st.container(border=True):
                    st.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
                    try:
                        st.image("img/MS-Logo.png", width=64)
                    except:
                        st.markdown("<h3>MS</h3>", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    ms_opciones = list(ms_creds_dict.keys())
                    if ms_opciones:
                        sel_ms = st.selectbox("Perfil SharePoint", ms_opciones, key="sel_ms_active", disabled=is_exec)
                        st.caption(f"Usuario: `{sel_ms}`")
                        
                        if st.button("🗑️ Eliminar MS", key="del_ms", disabled=is_exec, use_container_width=True):
                            if supabase:
                                sitio_val = f"MS_{st.session_state.get('logged_user', '')}_{sel_ms}"
                                try:
                                    supabase.table("credenciales_sitios").delete().eq("sitio", sitio_val).execute()
                                    del ms_creds_dict[sel_ms]
                                    st.session_state["ms_creds_dict"] = ms_creds_dict
                                    st.success("Eliminado")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {e}")
                    else:
                        st.warning("Sin configurar")
                        sel_ms = ""
        else:
            sel_ms = ""

        # --- FORMULARIO DE AGREGAR CREDENCIAL ---
        st.divider()
        if "mostrar_agregar_form" not in st.session_state:
            st.session_state.mostrar_agregar_form = False
            
        if not st.session_state.mostrar_agregar_form:
            if st.button("➕ Configurar Nueva Credencial", key="btn_show_add", use_container_width=True, disabled=is_exec):
                st.session_state.mostrar_agregar_form = True
                st.rerun()
        else:
            with st.container(border=True):
                c_form1, c_form2 = st.columns([1.2, 3])
                
                with c_form1:
                    sitio_opciones = ["Worklift (WL)", "Bureau Veritas (BV)"]
                    if show_ms:
                        sitio_opciones.append("Microsoft SharePoint (MS)")
                        
                    sitio_a_agregar = st.selectbox("Sitio Destino", sitio_opciones, key="add_sitio_sel")
                    
                    # Mostrar logo correspondiente dinámicamente
                    st.markdown("<div style='text-align: center; margin-top: 15px;'>", unsafe_allow_html=True)
                    if "Worklift" in sitio_a_agregar:
                        logo_form_path = "img/WL-Logo.png"
                    elif "Bureau" in sitio_a_agregar:
                        logo_form_path = "img/BV-Logo.png"
                    else:
                        logo_form_path = "img/MS-Logo.png"
                        
                    try:
                        st.image(logo_form_path, width=72)
                    except:
                        st.write("🔑")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                with c_form2:
                    st.markdown(f"#### Agregar Credencial para {sitio_a_agregar}")
                    new_u = st.text_input("Usuario / Correo", key="add_new_u")
                    new_p = st.text_input("Contraseña", type="password", key="add_new_p")
                    
                    c_btn1, c_btn2 = st.columns(2)
                    with c_btn1:
                        if st.button("💾 Guardar Credencial", key="btn_save_cred", use_container_width=True, type="primary"):
                            if not new_u or not new_p:
                                st.error("Ingresa usuario y contraseña")
                            else:
                                if "Worklift" in sitio_a_agregar:
                                    sitio_prefix = "WL"
                                    creds_dict = wl_creds_dict
                                    dict_key = "wl_creds_dict"
                                elif "Bureau" in sitio_a_agregar:
                                    sitio_prefix = "BV"
                                    creds_dict = bv_creds_dict
                                    dict_key = "bv_creds_dict"
                                else:
                                    sitio_prefix = "MS"
                                    creds_dict = ms_creds_dict
                                    dict_key = "ms_creds_dict"
                                    
                                if supabase:
                                    try:
                                        sitio_val = f"{sitio_prefix}_{st.session_state.get('logged_user', '')}_{new_u}"
                                        res_find = supabase.table("credenciales_sitios").select("id").eq("sitio", sitio_val).execute()
                                        if res_find.data:
                                            supabase.table("credenciales_sitios").update({"user_enc": encriptar(new_u), "pass_enc": encriptar(new_p)}).eq("sitio", sitio_val).execute()
                                        else:
                                            supabase.table("credenciales_sitios").insert({"usuario_app": st.session_state.get('logged_user', ''), "sitio": sitio_val, "user_enc": encriptar(new_u), "pass_enc": encriptar(new_p)}).execute()
                                        
                                        creds_dict[new_u] = new_p
                                        st.session_state[dict_key] = creds_dict
                                        st.success("Guardado correctamente!")
                                        st.session_state.mostrar_agregar_form = False
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error guardando en BD: {e}")
                    with c_btn2:
                        if st.button("❌ Cancelar", key="btn_cancel_add", use_container_width=True):
                            st.session_state.mostrar_agregar_form = False
                            st.rerun()

        # Synchronize selection with session state for bot execution
        st.session_state["wl_user"] = sel_wl if sel_wl else ""
        st.session_state["wl_pw"] = wl_creds_dict.get(sel_wl, "") if sel_wl else ""
        st.session_state["bv_user"] = sel_bv if sel_bv else ""
        st.session_state["bv_pw"] = bv_creds_dict.get(sel_bv, "") if sel_bv else ""
        st.session_state["ms_user"] = sel_ms if sel_ms else ""
        st.session_state["ms_pw"] = ms_creds_dict.get(sel_ms, "") if sel_ms else ""
    
    col_left, col_right = st.columns([1, 1.2], gap="large")
    
    with col_left:
        
        with st.container(border=True):
            if es_usuario_ingresos():
                c1, c2, c3 = st.columns(3)
                bajar_cert = c1.checkbox("Certificados", value=True, disabled=is_exec)
                bajar_inf = c2.checkbox("Informes", value=False, disabled=is_exec)
                bajar_doc_equipo = c3.checkbox("Doc. Equipo", value=False, disabled=is_exec, help="Descarga Título+Cédula o Factura de SharePoint")
            else:
                c1, c2 = st.columns(2)
                bajar_cert = c1.checkbox("Descargar Certificados", value=True, disabled=is_exec)
                bajar_inf = c2.checkbox("Descargar Informes", value=False, disabled=is_exec)
                bajar_doc_equipo = False
            es_semestral = st.checkbox("Vencimiento Semestral (180 días)", help="Calcula una alerta extra a los 6 meses.", disabled=is_exec)
            modo_pruebas = False
            if st.session_state.get("logged_user") == "fcendra":
                modo_pruebas = st.checkbox("🧪 Modo Pruebas (No inyecta métricas)", value=True, disabled=is_exec)
                modo_pruebas = st.checkbox("🧪 Modo Pruebas (No inyecta métricas)", value=True, disabled=is_exec)
                
        with st.expander("⚙️ Configuración de Salida", expanded=not is_exec):
            nombre_excel = st.text_input("Nombre del Excel", value="Reporte_Hopper", disabled=is_exec)
            nombre_zip = st.text_input("Nombre del Archivo ZIP", value="Certificados", disabled=is_exec)
            prefijo_cert = "" # Forzamos vacío para no pisar el nombre original de los PDFs
    
        st.markdown("##### Listado de Internos")
        archivo_subido = st.file_uploader("Subí tu Excel, TXT, CSV, PDF o Foto", type=['txt', 'csv', 'xlsx', 'png', 'jpg', 'jpeg', 'pdf'], help="También podés arrastrar el archivo.", disabled=is_exec)
        
        # --- LÓGICA DE COMPONENTE PORTAPAPELES (LIBRERÍA EXTERNA) ---
        if not is_exec:
            try:
                from streamlit_paste_button import paste_image_button
                paste_result = paste_image_button(
                    label="📋 Pegar Imagen del Portapapeles",
                    background_color="#008657",
                    hover_background_color="#006644",
                    key=f"paste_btn_{st.session_state.get('paste_key', 0)}"
                )
                if paste_result and paste_result.image_data is not None:
                    if not archivo_subido:
                        import io
                        buf = io.BytesIO()
                        paste_result.image_data.save(buf, format="PNG")
                        buf.name = "pasted_image.png"
                        buf.type = "image/png"
                        buf.size = len(buf.getvalue())
                        archivo_subido = buf
                        st.success("✅ Imagen pegada cargada correctamente.")
            except ImportError:
                pass

        if archivo_subido and archivo_subido.name.lower().endswith(('.png', '.jpg', '.jpeg', '.pdf')):
            st.info("⚠️ **Función Experimental:** La extracción de texto desde imagen (OCR) puede requerir revisión manual.")
            
        # LÓGICA DE EXTRACCIÓN AUTOMÁTICA
        if archivo_subido is not None:
            archivo_id = archivo_subido.name + str(archivo_subido.size)
            if st.session_state.ultimo_archivo_procesado != archivo_id:
                with st.spinner("Procesando archivo..."):
                    texto_extraido = ""
                    if archivo_subido.name.lower().endswith(('.png', '.jpg', '.jpeg')):
                        try:
                            # Cargar internos viejos para guiado de corrección
                            base_viejos = set()
                            ruta_viejos = "internos_viejos.txt"
                            if os.path.exists(ruta_viejos):
                                try:
                                    with open(ruta_viejos, "r", encoding="utf-8") as f:
                                        base_viejos = {line.strip().upper() for line in f if line.strip()}
                                except Exception as e:
                                    print(f"Error al leer internos_viejos.txt: {e}")

                            image = Image.open(archivo_subido).convert('L')
                            w, h = image.size
                            image = image.resize((w*2, h*2), Image.Resampling.LANCZOS)
                            enhancer = ImageEnhance.Contrast(image)
                            image = enhancer.enhance(2.0)
                            
                            texto_imagen = pytesseract.image_to_string(image, config='--psm 11')
                            palabras = texto_imagen.upper().split()
                            texto_corregido = []
                            for p in palabras:
                                for char in ['£', '€', 'È', 'É']: p = p.replace(char, 'E')
                                p = re.sub(r'[^A-Z0-9]', '', p)
                                if not p: continue
                                
                                # Si empieza con 3 y parece un interno nuevo, corregir a E
                                if p.startswith('3') and len(p) == 7:
                                    p = 'E' + p[1:]
                                # Si empieza con 4, ^, o @ y parece interno nuevo, corregir a A
                                elif p.startswith('3') and p[1:] in base_viejos:
                                    pass
                                elif p[0] in ['4', '^', '@'] and len(p) == 7:
                                    p = 'A' + p[1:]
                                    
                                if p.startswith('E') or p.startswith('A'):
                                    p_resto = p[1:].replace('O', '0').replace('I', '1').replace('L', '1').replace('S', '5').replace('Z', '2').replace('G', '6')
                                    texto_corregido.append(p[0] + p_resto)
                                else:
                                    # Para internos viejos:
                                    # Intentamos corrección puramente numérica
                                    p_corr_num = p.replace('O', '0').replace('I', '1').replace('L', '1').replace('S', '5').replace('Z', '2').replace('G', '6')
                                    if p_corr_num in base_viejos:
                                        texto_corregido.append(p_corr_num)
                                    # Si tiene letra final (B, C, D)
                                    elif len(p) > 1 and p[-1] in ['B', 'C', 'D']:
                                        p_cuerpo = p[:-1].replace('O', '0').replace('I', '1').replace('L', '1').replace('S', '5').replace('Z', '2').replace('G', '6')
                                        p_corr_letra = p_cuerpo + p[-1]
                                        if p_corr_letra in base_viejos:
                                            texto_corregido.append(p_corr_letra)
                                        else:
                                            texto_corregido.append(p_corr_num)
                                    else:
                                        texto_corregido.append(p_corr_num)
                                        
                            texto_extraido = " ".join(texto_corregido)
                        except Exception as e:
                            st.error(f"Error OCR: {e}")
                    else:
                        texto_extraido = extraer_texto_de_archivo(archivo_subido)
                    
                    lista_nuevos = extraer_internos(texto_extraido)
                    if lista_nuevos:
                        nuevo_texto = " ".join(lista_nuevos)
                        if st.session_state.texto_area:
                            st.session_state.texto_area += " " + nuevo_texto
                        else:
                            st.session_state.texto_area = nuevo_texto
                        st.success(f"✅ Se agregaron {len(lista_nuevos)} internos extraídos del archivo.")
                    else:
                        st.warning("No se detectaron internos válidos en el archivo.")
                
                st.session_state.ultimo_archivo_procesado = archivo_id
                st.rerun()
            
        texto_internos = st.text_area("Revisá o pegá los internos acá:", height=115, placeholder="E040230, 3797...", key="texto_area", disabled=is_exec)
        btn_run = st.button("🚀 COMENZAR PROCESO", use_container_width=True, disabled=is_exec)

    
    with col_right:
        terminal_placeholder = st.empty()
        def render_terminal():
            html = f'<div class="terminal-box">'
            html += f'<div style="font-family: \'Consolas\', monospace; font-weight: bold; font-size: 0.9rem; color: #00ff00; margin-bottom: 10px; border-bottom: 1px dashed #555; padding-bottom: 5px;">&gt;<span class="blink-cursor">_</span> REGISTRO DE ACTIVIDAD</div>'
            for entry in st.session_state.log_history:
                # Colores basados en íconos
                color = "#f8f9fa"
                if "✅" in entry or "VIGENTE" in entry: color = "#50fa7b"
                elif "❌" in entry or "ERROR" in entry or "VENCIDO" in entry: color = "#ff5555"
                elif "⚠️" in entry or "⏳" in entry or "PRÓXIMO" in entry: color = "#f1fa8c"
                elif "🤖" in entry: color = "#8be9fd"
                html += f'<div class="log-entry" style="color: {color};">{entry}</div>'
            html += '</div>'
            terminal_placeholder.markdown(html, unsafe_allow_html=True)
        render_terminal()
    
    if btn_run:
        wl_usr = st.session_state.get("wl_user", "")
        wl_pass = st.session_state.get("wl_pw", "")
        
        if not wl_usr or not wl_pass:
            st.error("Faltan credenciales de Worklift. Selecciona o guarda una.")
        elif bajar_doc_equipo and not st.session_state.get("ms_user", ""):
            st.error("Faltan credenciales de Microsoft SharePoint para descargar la documentación de equipos. Guarda una en la sección de credenciales.")
        else:
            # Almacenamos parámetros y disparamos ejecución diferida
            st.session_state.bajar_cert_val = bajar_cert
            st.session_state.bajar_inf_val = bajar_inf
            st.session_state.bajar_doc_equipo_val = bajar_doc_equipo
            st.session_state.es_semestral_val = es_semestral
            st.session_state.nombre_excel_val = nombre_excel
            st.session_state.nombre_zip_val = nombre_zip
            st.session_state.texto_internos_val = texto_internos
            st.session_state.ejecutando = True
            st.rerun()

    # --- FLUJO DE EJECUCIÓN DIFERIDO (BLOQUEO DE CONTROLES) ---
    if st.session_state.ejecutando:
        # Cargar parámetros almacenados
        bajar_cert = st.session_state.get("bajar_cert_val", True)
        bajar_inf = st.session_state.get("bajar_inf_val", False)
        bajar_doc_equipo = st.session_state.get("bajar_doc_equipo_val", False)
        es_semestral = st.session_state.get("es_semestral_val", False)
        nombre_excel = st.session_state.get("nombre_excel_val", "Reporte_Hopper")
        nombre_zip = st.session_state.get("nombre_zip_val", "Certificados")
        texto_internos = st.session_state.get("texto_internos_val", "")
        prefijo_cert = "" # Forzamos vacío para no pisar el nombre original de los PDFs
        
        wl_usr = st.session_state.get("wl_user", "")
        wl_pass = st.session_state.get("wl_pw", "")
        bv_usr = st.session_state.get("bv_user", "")
        bv_pw = st.session_state.get("bv_pw", "")
        ms_usr = st.session_state.get("ms_user", "")
        ms_pass = st.session_state.get("ms_pw", "")
        
        try:
            ruta_temp = "descargas_temp"
            if os.path.exists(ruta_temp): shutil.rmtree(ruta_temp)
            asegurar_carpeta(ruta_temp)
    
            st.session_state.proceso_completo = False
            
            # Formato de log inicial enriquecido
            si_no = lambda b: "Sí" if b else "No"
            st.session_state.log_history = [
                "--- PARÁMETROS ---",
                f"Descargar certificados: {si_no(bajar_cert)}",
                f"Descargar informes de inspección: {si_no(bajar_inf)}",
                f"Descargar doc. equipo (SharePoint): {si_no(bajar_doc_equipo)}",
                f"Modo vencimiento semestral: {si_no(es_semestral)}",
                "------------------",
                "Iniciando conexión con Worklift..."
            ]
            render_terminal()
            
            # --- AUTO SCROLL A LA TERMINAL ---
            st.components.v1.html("""<script>window.parent.document.querySelector('.terminal-box').scrollIntoView({behavior: 'smooth'});</script>""", height=0)
            
            # Solo se extrae del cuadro de texto final (ya combinado)
            lista = extraer_internos(texto_internos)
            
            if not lista:
                st.session_state.log_history.append("❌ No se encontraron internos para procesar en el cuadro de texto.")
                render_terminal()
            else:
                st.session_state.log_history.append("⏳ Iniciando sesión en Worklift<span class='loading-dots'></span>")
                render_terminal()
                
                bot = WLHopperBot(headless=True)
                if bot.iniciar(wl_usr, wl_pass):
                    st.session_state.log_history[-1] = "🔐 Login exitoso en Worklift."
                    
                    # Conexión inicial a Bureau Veritas si corresponde
                    exito_bv_login = False
                    if bv_usr and bv_pw:
                        st.session_state.log_history.append("Iniciando conexión con BV...")
                        render_terminal()
                        bv_test_bot = BureauVeritasBot(headless=True)
                        exito_bv_login, error_bv_login = bv_test_bot.iniciar(bv_usr, bv_pw, pw_instance=bot.pw)
                        bv_test_bot.cerrar()
                        if exito_bv_login:
                            st.session_state.log_history[-1] = "🔐 Login exitoso en BV."
                        else:
                            st.session_state.log_history[-1] = f"❌ Falló conexión con BV: {error_bv_login}"
                            
                    # Conexión inicial a Microsoft SharePoint si corresponde
                    exito_ms_login = False
                    if bajar_doc_equipo and ms_usr and ms_pass:
                        st.session_state.log_history.append("Iniciando conexión con Microsoft SharePoint...")
                        render_terminal()
                        ms_test_bot = MicrosoftSharePointBot(headless=True)
                        exito_ms_login, error_ms_login = ms_test_bot.iniciar(ms_usr, ms_pass, pw_instance=bot.pw)
                        ms_test_bot.cerrar()
                        if exito_ms_login:
                            st.session_state.log_history[-1] = "🔐 Login exitoso en SharePoint."
                        else:
                            st.session_state.log_history[-1] = f"❌ Falló conexión con SharePoint: {error_ms_login}"
                    
                    render_terminal()
                    
                    res_lista = []
                    for int_id in lista:
                        st.session_state.log_history.append(f"<br>--- Procesando {int_id} ---")
                        render_terminal()
                        res = bot.procesar_interno(int_id, ruta_temp, bajar_cert, bajar_inf, es_semestral=es_semestral, prefijo_cert=prefijo_cert)
                        res['id'] = int_id
                        res['proveedor'] = "Worklift"
                        
                        # --- AUDITORÍA DOBLE: BUREAU VERITAS ---
                        if bv_usr and bv_pw and exito_bv_login:
                            bv_bot = BureauVeritasBot(headless=True)
                            exito_bv, error_bv = bv_bot.iniciar(bv_usr, bv_pw, pw_instance=bot.pw)
                            
                            if exito_bv:
                                bv_res = bv_bot.procesar_interno(int_id, ruta_temp, bajar_cert=bajar_cert, bajar_inf=bajar_inf, prefijo_cert=prefijo_cert)
                                bv_bot.cerrar()
                                
                                # Si BV encontró algo útil, comparamos con WL
                                if bv_res.get('descargado') or bv_res.get('status') == 'VIGENTE (BV)' or bv_res.get('status') == 'Encontrado en BV':
                                    
                                    # Funciones helper para fechas
                                    def get_dt(d_str):
                                        if not d_str or d_str == "-": return datetime.min
                                        try: return datetime.strptime(d_str, "%d/%m/%Y")
                                        except: return datetime.min
                                        
                                    wl_v_dt = get_dt(res.get('venc_real') if 'venc_real' in res else res.get('venc'))
                                    bv_v_dt = get_dt(bv_res.get('venc'))
                                    
                                    wl_i_dt = get_dt(res.get('insp'))
                                    bv_i_dt = get_dt(bv_res.get('insp'))
                                    
                                    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                                    
                                    bv_tiene_cert_vigente = (bv_v_dt >= hoy)
                                    wl_tiene_cert_vigente = (wl_v_dt >= hoy)
                                    
                                    gana_bv_cert = False
                                    gana_bv_inf = False
                                    
                                    # DECISIÓN CERTIFICADO: Gana la fecha de vencimiento más lejana si es válida
                                    if bv_tiene_cert_vigente and (bv_v_dt > wl_v_dt):
                                        gana_bv_cert = True
                                        
                                    # DECISIÓN INFORME
                                    if gana_bv_cert:
                                        gana_bv_inf = True
                                    else:
                                        if bv_i_dt > wl_i_dt:
                                            gana_bv_inf = True
                                            
                                    if gana_bv_cert or gana_bv_inf:
                                        res['proveedor'] = "Bureau Veritas"
                                        res['cert'] = bv_res.get('cert', 'NO') if gana_bv_cert else res.get('cert', 'NO')
                                        res['inf'] = bv_res.get('informe', 'NO') if gana_bv_inf else res.get('inf', 'NO')
                                        
                                        if gana_bv_inf:
                                            res['insp'] = bv_res.get('insp', res.get('insp'))
                                            
                                        # Guardamos venc_real (el vencimiento anual de BV o el de WL)
                                        res['venc_real'] = bv_res.get('venc') if gana_bv_cert else res.get('venc_real', res.get('venc'))
                                        
                                        if es_semestral:
                                            # Vencimiento semestral es 180 días desde la inspección ganadora
                                            i_dt = get_dt(res['insp'])
                                            if i_dt != datetime.min:
                                                res['venc'] = (i_dt + timedelta(days=180)).strftime("%d/%m/%Y")
                                            else:
                                                res['venc'] = "-"
                                        else:
                                            if gana_bv_cert:
                                                res['venc'] = bv_res.get('venc', res.get('venc'))
                                        
                                        # Recalcular días restantes para status
                                        dias_restantes = (get_dt(res['venc']) - hoy).days if res['venc'] != "-" else -1
                                        
                                        if dias_restantes > 30:
                                            res['status'] = "VIGENTE"
                                            res['color'] = "VERDE"
                                            res['obs_final'] = f"{dias_restantes} días de vigencia."
                                            res['accion_final'] = "-"
                                        elif 0 <= dias_restantes <= 30:
                                            res['status'] = "PRÓXIMO A VENCER"
                                            res['color'] = "AMARILLO"
                                            res['obs_final'] = f"{dias_restantes} días de vigencia."
                                            res['accion_final'] = "Coordinar recertificación"
                                        else:
                                            res['status'] = "VENCIDO"
                                            res['color'] = "ROJO"
                                            if es_semestral:
                                                res['obs_final'] = f"Último certificado vencido en {res['venc']}." if res['venc'] != "-" else "Último certificado vencido."
                                            else:
                                                res['obs_final'] = "Último certificado vencido."
                                            obs_bv = bv_res.get('observaciones', '')
                                            if obs_bv and gana_bv_inf: res['obs_final'] += f"\nObservaciones BV: {obs_bv}"
                                            res['accion_final'] = "Coordinar recertificación urgente"
                                            
                                        res['log'] = [
                                            f"📄 Último Informe de Inspección: {res['insp']} (BV)",
                                            f"📅 Fecha vencimiento certificado: {res['venc']} (BV)"
                                        ]
                                        if dias_restantes <= 30:
                                            res['log'].append(f"💡 Sugerencia: {res['accion_final']}")
                                            
                                        # ELIMINAR ARCHIVOS PERDEDORES (WORKLIFT)
                                        archivos_wl = res.get("archivos_descargados", [])
                                        for f_path in archivos_wl:
                                            if os.path.exists(f_path):
                                                es_wl_cert = "Certificado" in f_path
                                                es_wl_inf = "Informe" in f_path
                                                if (es_wl_cert and gana_bv_cert) or (es_wl_inf and gana_bv_inf):
                                                    try: os.remove(f_path)
                                                    except: pass
                                                    
                                        # Eliminar archivos perdedores de BV
                                        archivos_bv = bv_res.get("archivos_descargados", [])
                                        for f_path in archivos_bv:
                                            if os.path.exists(f_path):
                                                es_bv_cert = "Certificado" in f_path
                                                es_bv_inf = "Informe" in f_path
                                                if (es_bv_cert and not gana_bv_cert) or (es_bv_inf and not gana_bv_inf):
                                                    try: os.remove(f_path)
                                                    except: pass
                                                    
                                    else:
                                        # Ganó WL. Borramos los que bajó BV
                                        archivos_bv = bv_res.get("archivos_descargados", [])
                                        for f_path in archivos_bv:
                                            if os.path.exists(f_path):
                                                try: os.remove(f_path)
                                                except: pass
                                
                        # --- DESCARGA DE DOCUMENTACIÓN DE EQUIPOS: SHAREPOINT ---
                        res['doc_equipo'] = "-"
                        res['doc_equipo_tipo'] = "-"
                        if bajar_doc_equipo and ms_usr and ms_pass and exito_ms_login:
                            st.session_state.log_history.append("Buscando en SharePoint...")
                            render_terminal()
                            ms_bot = MicrosoftSharePointBot(headless=True)
                            ms_init, ms_err = ms_bot.iniciar(ms_usr, ms_pass, pw_instance=bot.pw)
                            if ms_init:
                                ms_res = ms_bot.procesar_interno(int_id, ruta_temp, prefijo_cert=prefijo_cert)
                                ms_bot.cerrar()
                                if ms_res.get('descargado'):
                                    res['doc_equipo'] = ms_res.get('archivo')
                                    res['doc_equipo_tipo'] = ms_res.get('tipo_doc')
                                    st.session_state.log_history[-1] = f"✅ Encontrado en SharePoint ({ms_res.get('tipo_doc')})"
                                else:
                                    st.session_state.log_history[-1] = "❌ No se encontró en SharePoint."
                            else:
                                ms_bot.cerrar()
                                st.session_state.log_history[-1] = f"❌ Error SharePoint: {ms_err}"
                            render_terminal()

                        # --- IMPRESIÓN DEL LOG REESTRUCTURADO ---
                        st.session_state.log_history.append(f"Proveedor: {res.get('proveedor', 'Worklift')}")
                        
                        desc_inf = " (descargado)" if res.get('inf') == "SI" else ""
                        st.session_state.log_history.append(f"Última Inspección: {res.get('insp')}{desc_inf}")
                        
                        desc_cert = " (descargado)" if res.get('cert') == "SI" else ""
                        status_tag = f" ({res.get('status', 'VIGENTE')})"
                        st.session_state.log_history.append(f"Último Certificado: {res.get('venc_real', res.get('venc'))}{desc_cert}{status_tag}")
                        
                        if bajar_doc_equipo:
                            doc_status = f" ({res.get('doc_equipo_tipo')})" if res.get('doc_equipo') != "-" else ""
                            st.session_state.log_history.append(f"Doc. Equipo: {'SI' if res.get('doc_equipo') != '-' else 'NO'}{doc_status}")
                        
                        # Cálculo de vigencia
                        def get_dt(d_str):
                            if not d_str or d_str == "-": return datetime.min
                            try: return datetime.strptime(d_str, "%d/%m/%Y")
                            except: return datetime.min
                        v_dt = get_dt(res.get('venc'))
                        hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                        if v_dt != datetime.min:
                            dias = (v_dt - hoy).days
                            if dias >= 0:
                                dias_str = f"{dias} días de vigencia"
                            else:
                                dias_str = f"vencido hace {-dias} días"
                        else:
                            dias_str = "sin registro"
                        st.session_state.log_history.append(f"Días de vigencia: {dias_str}")
                        
                        if res.get('accion_final') != "-":
                            st.session_state.log_history.append(f"Comentarios: {res.get('accion_final')}")
                            
                        # Conservar advertencias OCR y sugerencias especiales
                        for log_line in res.get('log', []):
                            if any(log_line.strip().startswith(x) for x in ["⚠️", "🤖", "💡"]):
                                if log_line.strip().startswith("💡 Sugerencia:") and res.get('accion_final') != "-":
                                    continue
                                st.session_state.log_history.append(f"  {log_line.strip()}")
                                
                        res_lista.append(res)
                        
                        # Inyección de métricas (si no es prueba y no falló por completo)
                        if not modo_pruebas and res.get('status') != "No se pudo encontrar":
                            fue_exito = ("VERDE" in res.get('color', '') or "Descargado" in res.get('cert', ''))
                            registrar_metrica(int_id, "Archivo/Texto", exito=fue_exito)
                            
                        render_terminal()
        
                    bot.cerrar()
                    st.session_state.log_history.append("🔓 Sesiones cerradas correctamente.")
                    st.session_state.log_history.append("🏁 PROCESO FINALIZADO.")
                    st.session_state.res_lista = res_lista
                    st.session_state.proceso_completo = True
                    st.session_state.hay_archivos = len(os.listdir(ruta_temp)) > 0 if os.path.exists(ruta_temp) else False
                    st.session_state.paste_key = st.session_state.get('paste_key', 0) + 1 # Resetear portapapeles
        
                    # Generación de HTML
                    html = f"""<style>table {{ border-collapse: collapse; }} td {{ white-space: nowrap; text-align: center; vertical-align: middle; mso-number-format: "\\@"; padding: 5px 15px; }} th {{ white-space: nowrap; text-align: center; vertical-align: middle; padding: 10px 20px; }}</style>
                    <table id="hopperTable" width="100%" border="1" style="font-family: Calibri;">"""
                    
                    headers_html = ['INTERNO', 'PROVEEDOR', 'ESTADO', 'ÚLTIMA INSPECCIÓN']
                    if es_semestral:
                        headers_html += ['VENCIMIENTO SEMESTRAL', 'VENCIMIENTO REAL']
                    else:
                        headers_html += ['VENCIMIENTO<br>ÚLTIMO CERTIFICADO']
                    
                    if bajar_doc_equipo:
                        headers_html += ['DOC. EQUIPO']
                        
                    headers_html += ['CERTIFICADO', 'INFORME', 'OBSERVACIONES', 'ACCIONES']
                    
                    html += '<tr style="background-color: #008657; color: white; font-weight: bold;">'
                    for h_val in headers_html:
                        html += f'<th>{h_val}</th>'
                    html += '</tr>'
 
                    excel_data = []
                    for r in res_lista:
                        bg, tx, st_text = "#FFFFFF", "#000000", r['status'].upper()
                        cert_val = r['cert']
                        color_code = r.get('color', '').upper()
                        if color_code == "VERDE": bg, tx = "#C6EFCE", "#006100"
                        elif color_code == "AMARILLO": bg, tx = "#FFEB9C", "#9C5700"
                        elif color_code == "ROJO": bg, tx = "#FFC7CE", "#9C0006"
                        else:
                            if "VERDE" in st_text or "VIGENTE" in st_text or "APROBADO" in st_text: bg, tx = "#C6EFCE", "#006100"
                            elif "AMARILLO" in st_text or "PRÓXIMO" in st_text or "GESTIÓN" in st_text or "REINSPECCIONAR" in st_text: bg, tx = "#FFEB9C", "#9C5700"
                            elif "ROJO" in st_text or "VENCIDO" in st_text or "RECHAZADO" in st_text: bg, tx = "#FFC7CE", "#9C0006"
                        
                        html += f'<tr><td>{r["id"]}</td><td>{r.get("proveedor", "Worklift")}</td><td style="background-color: {bg}; color: {tx}; font-weight: bold;">{st_text}</td>'
                        if es_semestral:
                            html += f'<td>{r.get("insp", "N/A")}</td><td>{r.get("venc", "N/A")}</td><td>{r.get("venc_real", "N/A")}</td>'
                        else:
                            html += f'<td>{r.get("insp", "N/A")}</td><td>{r.get("venc", "N/A")}</td>'
                        
                        if bajar_doc_equipo:
                            doc_val = r.get('doc_equipo_tipo', '-')
                            html += f'<td style="font-weight: bold;">{doc_val}</td>'
                            
                        html += f'<td>{cert_val}</td><td>{r["inf"]}</td>'
                        html += f'<td style="text-align: left; max-width: 250px; white-space: normal;">{r.get("obs_final", "-")}</td>'
                        html += f'<td style="text-align: left; white-space: normal; padding-right: 30px;">{r.get("accion_final", "-")}</td></tr>'
                        
                        if es_semestral:
                            row_dict = {
                                "INTERNO": r["id"], 
                                "PROVEEDOR": r.get("proveedor", "Worklift"), 
                                "ESTADO": st_text, 
                                "ÚLTIMA INSPECCIÓN": r["insp"], 
                                "VENCIMIENTO SEMESTRAL": r["venc"],
                                "VENCIMIENTO REAL": r.get("venc_real", "-")
                            }
                        else:
                            row_dict = {
                                "INTERNO": r["id"], 
                                "PROVEEDOR": r.get("proveedor", "Worklift"), 
                                "ESTADO": st_text, 
                                "ÚLTIMA INSPECCIÓN": r["insp"], 
                                "VENCIMIENTO ÚLTIMO CERTIFICADO": r["venc"]
                            }
                        
                        if bajar_doc_equipo:
                            row_dict["DOC. EQUIPO"] = r.get("doc_equipo_tipo", "-")
                            
                        row_dict.update({
                            "CERTIFICADO": cert_val, "INFORME": r["inf"], 
                            "OBSERVACIONES": r.get("obs_final", "-"), "ACCIONES": r.get("accion_final", "-")
                        })
                        excel_data.append(row_dict)
                        
                    html += "</table>"
                    st.session_state.html_excel = html.replace("\n", "")
                    st.session_state.df_excel = pd.DataFrame(excel_data)
                    st.rerun()
                else:
                    st.session_state.log_history.append("❌ ERROR: Credenciales de Worklift incorrectas.")
                    render_terminal()
                    st.error("No se pudo iniciar sesión. Verificá tu usuario y contraseña de Worklift.")
        except Exception as ex:
            st.session_state.log_history.append(f"❌ ERROR INESPERADO: {ex}")
            render_terminal()
        finally:
            st.session_state.ejecutando = False
            st.rerun()
                
    st.divider()
    
    # Preparar el archivo Excel
    excel_buffer = BytesIO()
    if st.session_state.proceso_completo and st.session_state.df_excel is not None:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            st.session_state.df_excel.to_excel(writer, index=False, sheet_name='Reporte')
            
            # Formato Condicional y Ancho de Columnas
            worksheet = writer.sheets['Reporte']
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # Ajustar ancho de columnas y alinear
            col_letras = [col[0].column_letter for col in worksheet.columns]
            headers_list = [cell.value for cell in worksheet[1]]
            
            for column_cells in worksheet.columns:
                col_letra = column_cells[0].column_letter
                header_val = column_cells[0].value
                
                # Calcular longitud para auto-fit básico
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ancho_final = min(length + 2, 50)
                
                if header_val == "ESTADO":
                    ancho_final = max(ancho_final, 22) # Más ancho
                elif header_val == "DOC. EQUIPO":
                    ancho_final = 20
                elif header_val == "OBSERVACIONES":
                    ancho_final = 19.57
                elif header_val == "ACCIONES":
                    ancho_final = 27.43
                elif header_val == "INFORME":
                    ancho_final = 9.5
                
                worksheet.column_dimensions[col_letra].width = ancho_final
                
                # Aplicar centrado y wrap a todas las celdas
                for cell in column_cells:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            # Colores
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_font = Font(color="006100", bold=True)
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            yellow_font = Font(color="9C5700", bold=True)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            
            header_fill = PatternFill(start_color="008657", end_color="008657", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            if es_semestral:
                worksheet.insert_rows(1)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(worksheet[2]))
                title_cell = worksheet.cell(row=1, column=1)
                title_cell.value = "⚠️ REPORTE DE VENCIMIENTOS SEMESTRALES (180 DÍAS)"
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.fill = header_fill
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                header_row = 2
                worksheet.row_dimensions[1].height = 25
            else:
                header_row = 1
            
            for cell in worksheet[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            headers = [cell.value for cell in worksheet[header_row]]
            estado_idx = headers.index("ESTADO") + 1 if "ESTADO" in headers else -1
                
            for row in range(header_row + 1, worksheet.max_row + 1):
                res_idx = row - (header_row + 1)
                color_code = ""
                if res_idx >= 0 and res_idx < len(st.session_state.res_lista):
                    color_code = st.session_state.res_lista[res_idx].get('color', '').upper()
                    
                for idx in [estado_idx]:
                    if idx != -1:
                        cell = worksheet.cell(row=row, column=idx)
                        if cell.value:
                            val = str(cell.value).upper()
                            
                            # Si estamos en la columna de estado principal, usamos el color explícito si existe
                            if idx == estado_idx and color_code:
                                if color_code == "VERDE":
                                    cell.fill = green_fill
                                    cell.font = green_font
                                elif color_code == "AMARILLO":
                                    cell.fill = yellow_fill
                                    cell.font = yellow_font
                                elif color_code == "ROJO":
                                    cell.fill = red_fill
                                    cell.font = red_font
                            else:
                                # Fallback o para la columna de estado semestral
                                if "VERDE" in val or "VIGENTE" in val or "APROBADO" in val:
                                    cell.fill = green_fill
                                    cell.font = green_font
                                elif "AMARILLO" in val or "PRÓXIMO" in val or "GESTIÓN" in val or "REINSPECCIONAR" in val:
                                    cell.fill = yellow_fill
                                    cell.font = yellow_font
                                elif "ROJO" in val or "VENCIDO" in val or "RECHAZADO" in val or "ERROR" in val:
                                    cell.fill = red_fill
                                    cell.font = red_font


    excel_data = excel_buffer.getvalue()

    dcol1, dcol2, dcol3 = st.columns(3)
    
    with dcol1:
        if st.session_state.proceso_completo:
            components.html(f"""
                <div id="desktopBtnContainer" style="display: none; margin:0; padding:0; height: 45px; align-items: center;">
                    <button id="cBtn" style="width: 100%; height: 45px; background-color: {VERDE_SULLAIR}; color: white; border: none; border-radius: 4px; font-weight: bold; cursor: pointer; font-family: sans-serif; box-sizing: border-box;">
                        📋 Copiar Tabla Excel
                    </button>
                    <textarea id="hiddenTable" style="position:fixed; top:-1000px; opacity:0;">{st.session_state.html_excel}</textarea>
                </div>
                
                <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
                <div id="mobileBtnContainer" style="display: none; margin:0; padding:0; height: 45px; align-items: center;">
                    <button id="shareBtn" style="width: 100%; height: 45px; background-color: #25D366; color: white; border: none; border-radius: 4px; font-weight: bold; cursor: pointer; font-family: sans-serif; box-sizing: border-box;">
                        📱 Compartir Tabla como Imagen
                    </button>
                    <div id="captureArea" style="position: absolute; left: -9999px; background: white; padding: 10px;">
                        {st.session_state.html_excel}
                    </div>
                </div>

                <script>
                // Detección de dispositivo
                const isMobile = /Android|webOS|iPhone|iPad|iPod|BlackBerry|IEMobile|Opera Mini/i.test(navigator.userAgent);
                if (isMobile) {{
                    document.getElementById('mobileBtnContainer').style.display = 'flex';
                }} else {{
                    document.getElementById('desktopBtnContainer').style.display = 'flex';
                }}

                // Lógica de copiar (Desktop)
                document.getElementById('cBtn').onclick = function() {{
                    const btn = this;
                    const html = document.getElementById('hiddenTable').value;
                    const blob = new Blob([html], {{ type: 'text/html' }});
                    const data = [new ClipboardItem({{ 'text/html': blob }})];
                    navigator.clipboard.write(data).then(() => {{
                        btn.innerHTML = "✅ ¡COPIADO! (pegar con ctrl+v en Excel)";
                        btn.style.backgroundColor = "#28a745";
                        setTimeout(() => {{ btn.innerHTML = "📋 Copiar Tabla Excel"; btn.style.backgroundColor = "{VERDE_SULLAIR}"; }}, 2000);
                    }});
                }};

                // Lógica de compartir imagen (Móvil)
                document.getElementById('shareBtn').onclick = function() {{
                    const btn = this;
                    const originalText = btn.innerHTML;
                    btn.innerHTML = "⏳ Generando...";
                    
                    html2canvas(document.getElementById('captureArea')).then(canvas => {{
                        canvas.toBlob(blob => {{
                            const file = new File([blob], "reporte.png", {{ type: "image/png" }});
                            if (navigator.canShare && navigator.canShare({{ files: [file] }})) {{
                                navigator.share({{
                                    files: [file],
                                    title: 'Reporte WL Hopper',
                                    text: 'Reporte de Certificados'
                                }}).then(() => {{
                                    btn.innerHTML = "✅ Compartido";
                                    setTimeout(() => btn.innerHTML = originalText, 2000);
                                }}).catch(err => {{
                                    btn.innerHTML = "❌ Error al compartir";
                                    setTimeout(() => btn.innerHTML = originalText, 2000);
                                }});
                            }} else {{
                                const url = URL.createObjectURL(blob);
                                const a = document.createElement('a');
                                a.href = url;
                                a.download = "reporte.png";
                                a.click();
                                URL.revokeObjectURL(url);
                                btn.innerHTML = "✅ Descargado";
                                setTimeout(() => btn.innerHTML = originalText, 2000);
                            }}
                        }});
                    }});
                }};
                </script>
            """, height=45) 
        else:
            st.button("📋 Copiar Tabla / Imagen", disabled=True, use_container_width=True)
            
    with dcol2:
        if st.session_state.proceso_completo:
            safe_name = nombre_excel.strip() if nombre_excel.strip() else "Reporte_WLHopper"
            if not safe_name.endswith(".xlsx"): safe_name += ".xlsx"
            st.download_button("📊 Descargar Excel", data=excel_data, file_name=safe_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            st.button("📊 Descargar Excel", disabled=True, use_container_width=True)
    
    with dcol3:
        z_buf = BytesIO()
        if st.session_state.proceso_completo and st.session_state.hay_archivos:
            with zipfile.ZipFile(z_buf, "a", zipfile.ZIP_DEFLATED, False) as zf:
                for r, d, files in os.walk("descargas_temp"):
                    for f in files: zf.write(os.path.join(r, f), f)
        
        safe_zip = nombre_zip.strip() if nombre_zip.strip() else "certificados"
        if safe_zip.endswith(".xlsx"): safe_zip = safe_zip[:-5]
        if not safe_zip.endswith(".zip"): safe_zip += ".zip"
        
        st.download_button(
            "📂 Descargar Archivo ZIP", 
            data=z_buf.getvalue(), 
            file_name=safe_zip, 
            disabled=not (st.session_state.proceso_completo and st.session_state.hay_archivos), 
            use_container_width=True
        )
        
        # --- AUTO SCROLL A LOS BOTONES DE DESCARGA ---
        if st.session_state.proceso_completo:
            st.components.v1.html("""<script>
            setTimeout(() => {
                const buttons = window.parent.document.querySelectorAll('button[kind="primary"], button[kind="secondary"]');
                if (buttons.length > 0) {
                    buttons[buttons.length - 1].scrollIntoView({behavior: 'smooth'});
                }
            }, 500);
            </script>""", height=0)

    # El bloque de compartir como imagen se movió a dcol1 y se intercala por CSS.

    st.divider()
    st.caption("© 2026 - Desarrollado por Fede García Cendra para Sullair Argentina S.A.")
    st.caption("Consultas a: fcendra@sullair.com.ar")
